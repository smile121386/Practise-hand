import openpyxl


def read_test_data(sheet_name):
    wb = openpyxl.load_workbook('test_data.xlsx')
    ws = wb[sheet_name]
    list = []
    test_data = []
    for i in ws.rows:
        for row in i:
            list.append(row.value)
        test_data.append(list)
        list = []
    return test_data


def read_excel_data(excel_name, sheet_name, row_number='', column_number=''):
    wb = openpyxl.load_workbook(excel_name)
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_column = ws.max_column
    if row_number == '' and column_number == '':
        data = []
        for i in range(1, max_row + 1):
            for j in range(1, max_column + 1):
                data.append(ws.cell(row=i, column=j).value)
    elif row_number == '' and column_number != '':
        data = []
        for i in range(1, max_row + 1):
            data.append(ws.cell(row=i, column=column_number).value)
    elif row_number != '' and column_number == '':
        data = []
        for j in range(1, max_column + 1):
            data.append(ws.cell(row=row_number, column=j).value)
    else:
        data = ws.cell(row=row_number, column=column_number).value
    return data


def write_approver_data(approver):
    wb = openpyxl.load_workbook('test_data.xlsx')
    ws = wb['approver_list']
    ws.append(approver)
    wb.save('test_data.xlsx')


def clear_test_sheet(sheet):
    wb = openpyxl.load_workbook('test_data.xlsx')
    ws = wb[sheet]
    wb.remove(ws)
    wb.create_sheet(sheet)
    wb.save('test_data.xlsx')


def write_excel_data(excel_name, sheet_name):
    wb = openpyxl.load_workbook(excel_name)
    ws = wb[sheet_name]
    ws.append


def write_excel(excel_name, sheet_name, row_number, col_number, write_value):
    wb = openpyxl.load_workbook(excel_name)
    ws = wb[sheet_name]
    ws.cell(row=int(row_number), column=int(col_number), value=write_value)
    wb.save(excel_name)


write_excel('a.xlsx', 'a', '1', '2', '3')
