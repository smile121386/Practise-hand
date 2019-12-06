import openpyxl


wb = openpyxl.Workbook()
ws1 = wb.create_sheet('a')
wss = wb.sheetnames
print(wss[:])
ws2 = wb['a']

ac = wb['a']

ac.cell(row=1,column=1,value=1)

wb.save('a.xlsx')
