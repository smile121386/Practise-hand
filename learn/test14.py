import openpyxl
#
# def create_excel_for_sheet(file_name,sheet_names):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     wb.to_excel(sheet_names[0])
#     wb.save(file_name)
#
# a = ['q']
# b = ['w']
# create_excel_for_sheet('a.xlsx', a)
# create_excel_for_sheet('a.xlsx', b)

filename = 'a.xlsx'
wb = openpyxl.load_workbook(filename)
wb.create_sheet('c')
wb.save(filename)
