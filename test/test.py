from openpyxl import *


def read_excel(sheet, column=None, row=None):
    wb = load_workbook('test.xlsx')
    ws = wb[sheet]
    max_row = ws.max_row
    max_column = ws.max_column
    sheet_values = []
    letters = [chr(i) for i in range(65, 91)]
    if column is None and row is None:
        for rows in range(1, max_row+1):
            for columns in letters[:max_column]:
                value = ws['{ncolumn}{nrow}'.format(ncolumn=columns, nrow=rows)].value
                sheet_values.append(value)
    elif column is None and row:
        for columns in letters[:max_column]:
            value = ws['{ncolumn}{nrow}'.format(ncolumn=columns, nrow=row)].value
            sheet_values.append(value)
    elif column and row is None:
        for rows in range(1, max_row+1):
            value = ws['{ncolumn}{nrow}'.format(ncolumn=column, nrow=rows)].value
            sheet_values.append(value)
    else:
        value = ws['{ncolumn}{nrow}'.format(ncolumn=column, nrow=row)].value
        sheet_values.append(value)
    print(sheet_values)


def write_excel(sheet):
    wb = load_workbook('test.xlsx')
    ws = wb[sheet]
    max_row = ws.max_row
    max_column = ws.max_column

read_excel(sheet='Sheet2', column='A')
