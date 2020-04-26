import openpyxl

def read_test_data():
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb['Sheet1']
    list = []
    max_row = ws.max_row
    for rowNum in range(1, max_row+1):
        list.append(ws.cell(row=rowNum, column=1).value)
    return list


def write_approver_data(approver):
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb['Sheet3']
    ws.append(approver)
    wb.save('test.xlsx')


def write_excel(i, data):
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb['Sheet1']
    ws['B%s'%i] = data
    wb.save('test.xlsx')

# read_test_data()