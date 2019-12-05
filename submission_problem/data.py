import openpyxl

def title(filename):
    data = openpyxl.load_workbook(filename=filename)
    sheets = data.sheetnames
    sheet = data[sheets[0]]
    return sheet['A1'].value


def text(filename):
    data = openpyxl.load_workbook(filename=filename)
    sheets = data.sheetnames
    sheet = data[sheets[0]]
    return sheet['A2'].value

