import re
import openpyxl

with open('html.html', 'r', encoding='utf-8') as f:
    html = f.read()
pattern = re.compile('<tr(.*?)</tr>', re.S)
results = re.findall(pattern, html)
wb = openpyxl.Workbook('test.xlsx')
ws = wb.active
ws.title = "WorkSheetTitle"


del results[0]
j = 1
for i in results:

    p = re.compile('<td>(.*?)</td>', re.S)
    w = re.findall(p, i)
    print('学科：%s，成绩：%s' % (w[1], w[4]))
    c = ws.cell(row=1,column=j,value=w[1])
    d = ws.cell(row=2,column=j,value=w[4])
    j += 1


print(len(results))
