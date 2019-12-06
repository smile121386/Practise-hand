import re
import openpyxl

with open('html.html', 'r', encoding='utf-8') as f:
    html = f.read()
pattern = re.compile('<tr(.*?)</tr>', re.S)
results = re.findall(pattern, html)
wb = openpyxl.Workbook()
ws = wb.create_sheet('my_sheet')
# ws1 = wb.create_sheet('my_sheet')
sheet = wb['my_sheet']

sheet.cell(row=1, column=1, value='学科')
sheet.cell(row=1, column=2, value='成绩')
del results[0]
j = 0
b = 2
for i in results:
    p = re.compile('<td>(.*?)</td>', re.S)
    w = re.findall(p, i)
    # if w[2] == '任选':
    #     pass
    # else:
    # print('学科：%s，成绩：%s' % (w[4], w[8]))
    a = int(w[4]) * float(w[8])
    # print(a)
    j = j + a
    # print('j:%s' % j)
    sheet.cell(row=b, column=1, value=w[1])
    sheet.cell(row=b, column=2, value=w[4])
    b += 1

wb.save('test.xlsx')
# print(j)
